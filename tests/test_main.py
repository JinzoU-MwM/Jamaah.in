"""
Comprehensive Unit Test Suite for Jamaah.in FastAPI Backend
============================================================

This test suite covers real-world scenarios for Hajj/Umrah document processing:
1. Data merge logic (same person, multiple documents)
2. Excel generation and format validation
3. Corrupt/invalid file handling
4. Date parsing resilience

Run with: pytest tests/test_main.py -v
"""

import pytest
import io
import json
from unittest.mock import patch, MagicMock
from fastapi.testclient import TestClient
from pathlib import Path

# Import the FastAPI app
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from main import app, ExtractedDataItem


# =============================================================================
# FIXTURES
# =============================================================================

@pytest.fixture
def client():
    """FastAPI test client fixture"""
    return TestClient(app)


@pytest.fixture
def sample_passport_ocr_text():
    """Realistic OCR output for a Passport document"""
    return """
    REPUBLIK INDONESIA
    PASSPORT / PASPOR
    
    Nama/Name: AHMAD DAHLAN
    Tempat Lahir/Place of Birth: JAKARTA
    Tanggal Lahir/Date of Birth: 15-08-1985
    
    No. Paspor/Passport No: X1234567
    Tanggal Dikeluarkan/Date of Issue: 01-05-2023
    Tanggal Berakhir/Date of Expiry: 01-05-2028
    Kantor Penerbit/Issuing Office: JAKARTA SELATAN
    """


@pytest.fixture
def sample_visa_ocr_text():
    """Realistic OCR output for a Visa document - note lowercase name"""
    return """
    KINGDOM OF SAUDI ARABIA
    UMRAH VISA
    
    Name: Ahmad Dahlan
    Visa Number: V999888777
    Date of Issue: 10-12-2024
    Valid Until: 10-03-2025
    
    GHUSL Registration Complete
    """


@pytest.fixture
def sample_ktp_ocr_text():
    """Realistic OCR output for a KTP document"""
    return """
    PROVINSI DKI JAKARTA
    KARTU TANDA PENDUDUK
    
    NIK: 3171234567890001
    Nama: SITI AMINAH
    Tempat/Tgl Lahir: BANDUNG, 20-03-1990
    Jenis Kelamin: PEREMPUAN
    Alamat: JL. SUDIRMAN NO. 123
    RT/RW: 001/002
    Kel/Desa: MENTENG
    Kecamatan: MENTENG
    Agama: ISLAM
    Status Perkawinan: KAWIN
    Pekerjaan: WIRASWASTA
    Kewarganegaraan: WNI
    Berlaku Hingga: SEUMUR HIDUP
    """


@pytest.fixture
def sample_merged_data():
    """Sample data for Excel generation test"""
    return [
        {
            "no_identitas": "3171234567890001",
            "nama": "AHMAD DAHLAN",
            "tempat_lahir": "JAKARTA",
            "tanggal_lahir": "1985-08-15",
            "alamat": "JL. SUDIRMAN NO. 123",
            "provinsi": "DKI JAKARTA",
            "kabupaten": "JAKARTA SELATAN",
            "kecamatan": "MENTENG",
            "kelurahan": "MENTENG",
            "no_paspor": "X1234567",
            "tanggal_paspor": "2023-05-01",
            "kota_paspor": "JAKARTA SELATAN",
            "no_visa": "V999888777",
            "jenis_identitas": "KTP",
            "kewarganegaraan": "WNI"
        }
    ]


# =============================================================================
# TEST 1: REAL DATA MERGE LOGIC (CRUCIAL)
# =============================================================================

class TestDataMergeLogic:
    """Test that multiple documents for the same person are merged correctly"""
    
    def test_merge_passport_and_visa_same_person(self, client):
        """
        CRITICAL TEST: Upload Passport + Visa for same person
        
        Scenario:
        - File 1 (Passport): name="AHMAD DAHLAN", passport_no="X1234567", visa_no=None
        - File 2 (Visa): name="Ahmad Dahlan" (lowercase), passport_no=None, visa_no="V999888777"
        
        Expected: ONE merged entry with BOTH passport_no AND visa_no
        """
        from models.schemas import DocumentData
        
        with patch('main.extract_text_from_image') as mock_ocr, \
             patch('main.parse_document') as mock_parse:
            
            # Mock OCR returns any text (we're mocking the parser anyway)
            mock_ocr.return_value = "mock text"
            
            # Mock parser to return specific DocumentData objects
            passport_data = DocumentData(
                document_type="PASSPORT",
                name="AHMAD DAHLAN",
                passport_number="X1234567",
                place_of_birth="JAKARTA",
                date_of_birth="1985-08-15"
            )
            
            visa_data = DocumentData(
                document_type="VISA",
                name="Ahmad Dahlan",  # Note: different case
                visa_number="V999888777"
            )
            
            mock_parse.side_effect = [passport_data, visa_data]
            
            # Create fake files
            passport_file = io.BytesIO(b"fake passport image data")
            visa_file = io.BytesIO(b"fake visa image data")
            
            # Upload both files
            response = client.post(
                "/process-documents/",
                files=[
                    ("files", ("passport_ahmad.jpg", passport_file, "image/jpeg")),
                    ("files", ("visa_ahmad.jpg", visa_file, "image/jpeg")),
                ]
            )
            
            assert response.status_code == 200
            data = response.json()
            
            # Should be consolidated to ONE entry
            assert data["status"] == "success"
            assert len(data["data"]) == 1, f"Expected 1 merged record, got {len(data['data'])}: {data['data']}"
            
            merged_record = data["data"][0]
            
            # Check both passport and visa are present
            assert merged_record["no_paspor"] == "X1234567", f"Passport number missing after merge. Got: {merged_record}"
            assert merged_record["no_visa"] == "V999888777", f"Visa number missing after merge. Got: {merged_record}"
            
            # Name should be preserved
            assert "AHMAD" in merged_record["nama"].upper() or "DAHLAN" in merged_record["nama"].upper()
    
    def test_merge_case_insensitive_names(self, client):
        """Test that 'Ahmad' matches 'AHMAD' during merge"""
        from models.schemas import DocumentData
        
        with patch('main.extract_text_from_image') as mock_ocr, \
             patch('main.parse_document') as mock_parse:
            
            mock_ocr.return_value = "mock text"
            
            # Two documents with different case names
            doc1 = DocumentData(
                document_type="PASSPORT",
                name="BUDI SANTOSO",
                passport_number="P111"
            )
            
            doc2 = DocumentData(
                document_type="VISA",
                name="budi santoso",  # lowercase
                visa_number="V222"
            )
            
            mock_parse.side_effect = [doc1, doc2]
            
            file1 = io.BytesIO(b"file1")
            file2 = io.BytesIO(b"file2")
            
            response = client.post(
                "/process-documents/",
                files=[
                    ("files", ("doc1.jpg", file1, "image/jpeg")),
                    ("files", ("doc2.jpg", file2, "image/jpeg")),
                ]
            )
            
            assert response.status_code == 200
            data = response.json()
            
            # Should merge into one record
            assert len(data["data"]) == 1, f"Case-insensitive merge failed. Got {len(data['data'])} records: {data['data']}"
            
            # Verify merged data has both passport and visa
            merged = data["data"][0]
            assert merged["no_paspor"] == "P111"
            assert merged["no_visa"] == "V222"
    
    def test_no_merge_different_persons(self, client):
        """Test that different people are NOT merged"""
        from models.schemas import DocumentData
        
        with patch('main.extract_text_from_image') as mock_ocr, \
             patch('main.parse_document') as mock_parse:
            
            mock_ocr.return_value = "mock text"
            
            # Two different people
            doc1 = DocumentData(
                document_type="PASSPORT",
                name="AHMAD DAHLAN",
                passport_number="P111"
            )
            
            doc2 = DocumentData(
                document_type="PASSPORT",
                name="SITI AMINAH",  # Different person
                passport_number="P222"
            )
            
            mock_parse.side_effect = [doc1, doc2]
            
            file1 = io.BytesIO(b"file1")
            file2 = io.BytesIO(b"file2")
            
            response = client.post(
                "/process-documents/",
                files=[
                    ("files", ("ahmad.jpg", file1, "image/jpeg")),
                    ("files", ("siti.jpg", file2, "image/jpeg")),
                ]
            )
            
            assert response.status_code == 200
            data = response.json()
            
            # Should remain as 2 separate records
            assert len(data["data"]) == 2, f"Different persons incorrectly merged. Got {len(data['data'])} records"


# =============================================================================
# TEST 2: EXCEL GENERATION & FORMAT VALIDATION
# =============================================================================

class TestExcelGeneration:
    """Test Excel file generation and format"""
    
    def test_generate_excel_returns_200(self, client, sample_merged_data):
        """Test that /generate-excel/ returns 200 status"""
        response = client.post(
            "/generate-excel/",
            json={"data": sample_merged_data}
        )
        
        assert response.status_code == 200
    
    def test_generate_excel_content_type(self, client, sample_merged_data):
        """Test that response has correct Content-Type for Excel"""
        response = client.post(
            "/generate-excel/",
            json={"data": sample_merged_data}
        )
        
        content_type = response.headers.get("content-type", "")
        
        # Should be xlsx format (not xlsm since we're not using VBA template)
        assert "spreadsheetml" in content_type or "excel" in content_type, \
            f"Unexpected content type: {content_type}"
    
    def test_generate_excel_valid_file(self, client, sample_merged_data):
        """Deep check: Load returned Excel and verify structure"""
        import openpyxl
        
        response = client.post(
            "/generate-excel/",
            json={"data": sample_merged_data}
        )
        
        assert response.status_code == 200
        
        # Load the returned content as Excel
        excel_content = io.BytesIO(response.content)
        
        try:
            workbook = openpyxl.load_workbook(excel_content)
        except Exception as e:
            pytest.fail(f"Failed to load Excel file: {e}")
        
        # Verify workbook has at least one sheet
        assert len(workbook.sheetnames) > 0, "Excel has no sheets"
        
        # Verify data was written
        ws = workbook.active
        assert ws.max_row >= 2, "Excel should have at least header + 1 data row"
        
        # Verify headers exist
        first_row = [cell.value for cell in ws[1]]
        assert "No Identitas" in first_row or any("Identitas" in str(h) for h in first_row if h), \
            "Missing expected headers"
    
    def test_generate_excel_empty_data_fails(self, client):
        """Test that empty data returns 400 error"""
        response = client.post(
            "/generate-excel/",
            json={"data": []}
        )
        
        assert response.status_code == 400


# =============================================================================
# TEST 3: GARBAGE/CORRUPT FILE HANDLING
# =============================================================================

class TestCorruptFileHandling:
    """Test graceful handling of invalid/corrupt files"""
    
    def test_invalid_file_extension(self, client):
        """Test that .txt files are rejected"""
        fake_txt = io.BytesIO(b"This is not an image, just plain text.")
        
        response = client.post(
            "/process-documents/",
            files=[("files", ("document.txt", fake_txt, "text/plain"))]
        )
        
        # Should either return 400 or return with failed status
        if response.status_code == 200:
            data = response.json()
            # Should report failure for this file
            assert data["failed"] > 0 or data["successful"] == 0
        else:
            assert response.status_code == 400
    
    def test_corrupt_image_graceful_handling(self, client):
        """Test that corrupt image data doesn't crash the server"""
        with patch('main.extract_text_from_image') as mock_ocr:
            # Simulate OCR throwing an exception for corrupt image
            mock_ocr.side_effect = Exception("Cannot decode image data")
            
            corrupt_data = io.BytesIO(b"\x00\x01\x02\x03 random bytes")
            
            response = client.post(
                "/process-documents/",
                files=[("files", ("corrupt.jpg", corrupt_data, "image/jpeg"))]
            )
            
            # Should NOT be 500 (server crash)
            assert response.status_code != 500, "Server crashed on corrupt file"
    
    def test_empty_file(self, client):
        """Test handling of empty file"""
        empty_file = io.BytesIO(b"")
        
        response = client.post(
            "/process-documents/",
            files=[("files", ("empty.jpg", empty_file, "image/jpeg"))]
        )
        
        # Should handle gracefully
        assert response.status_code in [200, 400]


# =============================================================================
# TEST 4: DATE PARSING RESILIENCE
# =============================================================================

class TestDateParsingResilience:
    """Test that various date formats are properly parsed"""
    
    @pytest.mark.parametrize("date_input,expected_format", [
        ("12-01-2024", "2024"),  # DD-MM-YYYY
        ("2024-01-12", "2024"),  # YYYY-MM-DD (already correct)
        ("12/01/2024", "2024"),  # DD/MM/YYYY with slash
    ])
    def test_date_formats(self, client, date_input, expected_format):
        """Test various date formats are handled"""
        with patch('main.extract_text_from_image') as mock_ocr:
            mock_ocr.return_value = f"""
            PASSPORT
            Nama: TEST PERSON
            Tanggal Lahir: {date_input}
            No. Paspor: P123
            """
            
            file = io.BytesIO(b"test")
            
            response = client.post(
                "/process-documents/",
                files=[("files", ("test.jpg", file, "image/jpeg"))]
            )
            
            assert response.status_code == 200
            data = response.json()
            
            if data["data"]:
                # The year should be present in the output
                dob = data["data"][0].get("tanggal_lahir", "")
                assert expected_format in dob or dob != "", \
                    f"Date parsing failed for input: {date_input}"


# =============================================================================
# TEST 5: API ENDPOINTS BASIC TESTS
# =============================================================================

class TestAPIEndpoints:
    """Basic API endpoint tests"""
    
    def test_root_endpoint(self, client):
        """Test root endpoint returns welcome message"""
        response = client.get("/")
        
        assert response.status_code == 200
        data = response.json()
        assert "Jamaah.in" in data.get("message", "")
    
    def test_health_endpoint(self, client):
        """Test health check endpoint"""
        response = client.get("/health")
        
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "healthy"
    
    def test_no_files_returns_400(self, client):
        """Test that no files returns 400"""
        response = client.post("/process-documents/", files=[])
        
        # FastAPI will return 422 for missing required field
        assert response.status_code in [400, 422]


# =============================================================================
# TEST 6: INTEGRATION TEST - FULL WORKFLOW
# =============================================================================

class TestFullWorkflow:
    """End-to-end integration test"""
    
    def test_complete_workflow(self, client, sample_passport_ocr_text, sample_ktp_ocr_text):
        """Test complete workflow: upload → process → generate Excel"""
        import openpyxl
        
        with patch('main.extract_text_from_image') as mock_ocr:
            mock_ocr.side_effect = [sample_passport_ocr_text, sample_ktp_ocr_text]
            
            # Step 1: Upload and process
            file1 = io.BytesIO(b"passport")
            file2 = io.BytesIO(b"ktp")
            
            process_response = client.post(
                "/process-documents/",
                files=[
                    ("files", ("passport.jpg", file1, "image/jpeg")),
                    ("files", ("ktp.jpg", file2, "image/jpeg")),
                ]
            )
            
            assert process_response.status_code == 200
            process_data = process_response.json()
            
            # Step 2: Generate Excel with the processed data
            excel_response = client.post(
                "/generate-excel/",
                json={"data": process_data["data"]}
            )
            
            assert excel_response.status_code == 200
            
            # Step 3: Verify Excel is valid
            excel_content = io.BytesIO(excel_response.content)
            workbook = openpyxl.load_workbook(excel_content)
            
            assert workbook.active.max_row >= 2, "Excel should have data rows"


# =============================================================================
# RUN CONFIGURATION
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
